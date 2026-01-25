#excel_data_handler.py
import openpyxl
from collections import defaultdict

def cargar_datos_excel(archivo='datos_produccion.xlsx'):
    """
    Carga inicial del archivo Excel y retorna el libro de trabajo
    para ser procesado por área
    """
    try:
        wb = openpyxl.load_workbook(archivo)
        return wb
    except Exception as e:
        print(f"Error al cargar el archivo Excel: {e}")
        raise

def obtener_areas_disponibles(ws):
    """
    Obtiene la lista de áreas únicas desde la hoja de líneas de producción.
    
    Args:
        ws: Worksheet de openpyxl (hoja Lineas_Produccion)
    
    Returns:
        list: Lista ordenada de áreas únicas
    
    Ejemplo:
        areas = obtener_areas_disponibles(wb['Lineas_Produccion'])
        # Retorna: ['ICT', 'SMT']
    """
    areas = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Primera columna es Area
            areas.add(row[0])
            print(f"Área encontrada: {row[0]}")
    
    areas_ordenadas = sorted(list(areas))
    print(f"Total de áreas identificadas: {len(areas_ordenadas)}")
    print(f"Áreas disponibles: {areas_ordenadas}")
    
    return areas_ordenadas

def cargar_lineas_produccion(ws):
    lineas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        lineas.append({
            'nombre': row[1],        # Ahora la columna 1 es el nombre
            'tiempo_disponible': row[2],
            'eficiencia': row[3]
        })
    return lineas

def cargar_lineas_produccion_area(ws, area):
    """Carga solo las líneas del área especificada"""
    lineas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == area:  # Primera columna es Area
            lineas.append({
                'nombre': row[1],
                'tiempo_disponible': row[2],
                'eficiencia': row[3],
                'area': row[0]  # Añadimos el área al diccionario
            })
    print(f"Cargadas {len(lineas)} líneas para el área {area}")
    return lineas

def cargar_modelos(ws):
    modelos = defaultdict(lambda: {'pcbs': defaultdict(dict)})
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Nuevos índices que coinciden con el Excel
            bu = row[0]                     # Columna A
            area_modelo = row[1]            # Columna B
            familia = row[2]                # Columna C
            nombre = row[3]                 # Columna D
            lado = row[4]                   # Columna E
            tiempo_ciclo = row[5]           # Columna F
            cantidad_por_producto = row[6]   # Columna G
            lineas_compatibles = row[7]     # Columna H
            prioridad = row[8]              # Columna I
            eficiencia = row[9]             # Columna J
            
            print(f"Procesando fila {index}: Familia={familia}, Nombre={nombre}")
            
            if familia not in modelos:
                modelos[familia] = {
                    'bu': bu,
                    'area': area_modelo,
                    'nombre': nombre,
                    'lineas_compatibles': [],
                    'prioridad': prioridad,
                    'eficiencia': eficiencia,
                    'pcbs': defaultdict(dict)
                }
            
            if lineas_compatibles:
                modelos[familia]['lineas_compatibles'] = [line.strip() for line in str(lineas_compatibles).split(',') if line.strip()]
                print(f"  Líneas compatibles: {modelos[familia]['lineas_compatibles']}")
            else:
                print(f"  Advertencia: No hay líneas compatibles para {familia}")
            
            if lado and tiempo_ciclo is not None and cantidad_por_producto is not None:
                modelos[familia]['pcbs'][nombre][lado] = {
                    'tiempo_ciclo': tiempo_ciclo,
                    'cantidad_por_producto': cantidad_por_producto
                }
                print(f"  PCB añadido: {nombre} - {lado}")
            else:
                print(f"  Advertencia: Datos de PCB incompletos para {familia} - {nombre}")
        
        except Exception as e:
            print(f"Error procesando la fila {index}: {e}")
            print(f"Contenido de la fila: {row}")
    
    return dict(modelos)

def cargar_modelos_area(ws, area):
    """Carga solo los modelos que pertenecen al área especificada"""
    modelos = defaultdict(lambda: {'pcbs': defaultdict(dict)})
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Nuevos índices que coinciden con el Excel
            bu = row[0]                     # Columna A
            area_modelo = row[1]            # Columna B
            familia = row[2]                # Columna C
            nombre = row[3]                 # Columna D
            lado = row[4]                   # Columna E
            tiempo_ciclo = row[5]           # Columna F
            cantidad_por_producto = row[6]   # Columna G
            lineas_compatibles = row[7]     # Columna H
            prioridad = row[8]              # Columna I
            eficiencia = row[9]             # Columna J
            
            if area_modelo == area:  # Solo procesar si corresponde al área
                if familia not in modelos:
                    modelos[familia] = {
                        'bu': bu,
                        'area': area_modelo,
                        'nombre': nombre,
                        'lineas_compatibles': [line.strip() for line in str(lineas_compatibles).split(',') if line.strip()],
                        'prioridad': prioridad,
                        'eficiencia': eficiencia,
                        'pcbs': defaultdict(dict)
                    }
                
                if lado and tiempo_ciclo is not None and cantidad_por_producto is not None:
                    modelos[familia]['pcbs'][nombre][lado] = {
                        'tiempo_ciclo': tiempo_ciclo,
                        'cantidad_por_producto': cantidad_por_producto
                    }
                    print(f"  PCB añadido para {area}: {nombre} - {lado}")
        except Exception as e:
            print(f"Error procesando la fila {index} para área {area}: {e}")
            print(f"Contenido de la fila: {row}")
    
    print(f"Cargados {len(modelos)} modelos para el área {area}")
    return dict(modelos)

def cargar_volumenes_produccion(ws):
    volumenes = {
        'anos_produccion': [cell.value for cell in ws[1][2:]],
        'modelos': []
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        volumenes['modelos'].append({
            'Familia': row[0],
            'dias_operacion_anual': row[1],
            'volumenes_anuales': list(row[2:])
        })
    return volumenes

def imprimir_resumen_volumenes(volumenes):
    print("\nResumen de volúmenes anuales por familia:")
    for ano_index, ano in enumerate(volumenes['anos_produccion']):
        print(f"\nAño {ano}:")
        for modelo in volumenes['modelos']:
            if ano_index < len(modelo['volumenes_anuales']):
                volumen = modelo['volumenes_anuales'][ano_index]
                print(f"  Familia {modelo['Familia']}: {volumen}")
            else:
                print(f"  Familia {modelo['Familia']}: No hay datos para este año")

def escribir_resultados_excel(df_resultados, df_resumen, archivo='datos_produccion.xlsx', nombre_hoja='Resultados'):
    """
    Escribe los resultados y resumen en una hoja específica del archivo Excel.
    Parámetros:
        df_resultados: DataFrame con los resultados detallados
        df_resumen: DataFrame con el resumen
        archivo: Nombre del archivo Excel
        nombre_hoja: Nombre de la hoja donde se escribirán los resultados (ej: 'Resultados_ICT')
    """
    try:
        # Intentamos abrir el archivo existente
        wb = openpyxl.load_workbook(archivo)
    except FileNotFoundError:
        # Si el archivo no existe, creamos uno nuevo
        wb = openpyxl.Workbook()
    
    # Si la hoja existe, la eliminamos para crear una nueva
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]
    
    # Creamos una nueva hoja
    ws = wb.create_sheet(nombre_hoja)

    # Escribir encabezados de la tabla principal
    headers = list(df_resultados.columns)
    ws.append(headers)

    # Escribir datos de la tabla principal
    for _, row in df_resultados.iterrows():
        ws.append(row.tolist())

    # Agregar dos filas en blanco
    ws.append([])
    ws.append([])

    # Escribir el título de la tabla de resumen
    ws.append(["Tabla de Resumen"])

    # Escribir encabezados de la tabla de resumen
    resumen_headers = list(df_resumen.columns)
    ws.append(resumen_headers)

    # Escribir datos de la tabla de resumen
    for _, row in df_resumen.iterrows():
        ws.append(row.tolist())

    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardamos el archivo
    wb.save(archivo)
    print(f"Resultados guardados en {archivo}, hoja '{nombre_hoja}'")